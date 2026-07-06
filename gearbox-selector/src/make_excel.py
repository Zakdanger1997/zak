#!/usr/bin/env python3
"""Bilingual (TR/EN) workbook of every per-application I-MAK formula, built from the
same metadata the app uses (apps.json), so it can't drift from the code."""
import json, pathlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = pathlib.Path(__file__).parent
APPS = json.load(open(HERE/"apps.json"))
OUT  = "/home/user/zak/gearbox-selector/IMAK-Reduktor-Formuller.xlsx"

HEAD="262B32"; RED="E11D13"
hdr=PatternFill("solid",fgColor=HEAD); redf=PatternFill("solid",fgColor=RED); altf=PatternFill("solid",fgColor="EEF1F5")
wf=Font(color="FFFFFF",bold=True); tf=Font(color="FFFFFF",bold=True,size=14); bold=Font(bold=True)
thin=Side(style="thin",color="C9CED6"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(wrap_text=True,vertical="top"); topa=Alignment(vertical="top")

def banner(ws,text,nc):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=nc)
    c=ws.cell(row=1,column=1,value=text); c.fill=redf; c.font=tf
    c.alignment=Alignment(vertical="center",horizontal="left",indent=1); ws.row_dimensions[1].height=30
def head(ws,row,nc):
    for c in range(1,nc+1):
        x=ws.cell(row=row,column=c); x.fill=hdr; x.font=wf; x.border=bd
        x.alignment=Alignment(wrap_text=True,vertical="center")
def put(ws,vals,zebra=True,boldcols=()):
    ws.append(vals); r=ws.max_row
    for c in range(1,len(vals)+1):
        x=ws.cell(row=r,column=c); x.border=bd; x.alignment=wrap
        if zebra and r%2==0: x.fill=altf
        if c in boldcols: x.font=bold
    return r

wb=openpyxl.Workbook()

# ---- Sheet 1: Applications ----
ws=wb.active; ws.title="Uygulamalar-Applications"
cols=["#","Grup / Group (TR)","Group (EN)","Uygulama (TR)","Application (EN)","Formül (TR)","Formula (EN)",
      "Sf","η","Önerilen ürün (TR)","Suggested product (EN)","Örnek Mᴛ (Nm)","Örnek n (d/dk)","Örnek motor (kW)"]
banner(ws,"İ-MAK REDÜKTÖR — Uygulamaya özel formüller / Per-application formulas",len(cols))
ws.append(cols); head(ws,2,len(cols))
for i,a in enumerate(APPS,1):
    ex=a.get("ex",{})
    put(ws,[i,a["grp"]["tr"],a["grp"]["en"],a["name"]["tr"],a["name"]["en"],a["formula"]["tr"],a["formula"]["en"],
        a["sf"],a["eta"],a["series"]["tr"],a["series"]["en"],ex.get("MT"),ex.get("n"),ex.get("Pmotor")],boldcols=(4,5))
for i,w in enumerate([4,20,20,30,30,34,34,6,6,26,26,13,13,14],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="A3"

# ---- Sheet 2: Variables ----
ws2=wb.create_sheet("Degiskenler-Variables")
cols2=["Uygulama (TR)","Application (EN)","Değişken / Variable","Etiket (TR)","Label (EN)","Birim / Unit","Varsayılan / Default"]
banner(ws2,"Değişkenler ve varsayılan değerler / Variables & default values",len(cols2))
ws2.append(cols2); head(ws2,2,len(cols2))
for a in APPS:
    for v in a["vars"]:
        put(ws2,[a["name"]["tr"],a["name"]["en"],v["id"],v["tr"],v["en"],v["u"],v["def"]],boldcols=(3,))
for i,w in enumerate([30,30,16,28,28,14,16],1): ws2.column_dimensions[get_column_letter(i)].width=w
ws2.freeze_panes="A3"

# ---- Sheet 3: Worked examples ----
ws3=wb.create_sheet("Ornek Hesap-Worked example")
cols3=["Uygulama (TR)","Application (EN)","Adım (TR)","Step (EN)","İfade / Expression","Sonuç / Result","Birim / Unit"]
banner(ws3,"Adım adım örnek hesap (varsayılan değerlerle) / Step-by-step worked example",len(cols3))
ws3.append(cols3); head(ws3,2,len(cols3))
for a in APPS:
    steps=a.get("ex",{}).get("steps",[])
    for j,s in enumerate(steps):
        put(ws3,[a["name"]["tr"] if j==0 else "", a["name"]["en"] if j==0 else "", s["tr"],s["en"],s["expr"],s["res"],s["u"]],boldcols=(1,))
for i,w in enumerate([28,28,26,26,40,14,12],1): ws3.column_dimensions[get_column_letter(i)].width=w
ws3.freeze_panes="A3"

# ---- Sheet 4: Constants ----
ws4=wb.create_sheet("Sabitler-Constants")
banner(ws4,"Sabitler & servis faktörü / Constants & service factor",3)
ws4.append(["Öğe / Item","Değer / Value","Açıklama / Explanation"]); head(ws4,2,3)
rows=[
 ("g","9,81 m/s²","Yerçekimi ivmesi / Gravity."),
 ("Moment sabiti / Moment const.","9550","Mᴛ[Nm] = 9550·P[kW]·η/n[d/dk]  ·  (=60000/2π)."),
 ("Konveyör sabiti / Conveyor const.","367","Helezon & elevatör; = 3600·g/9,81."),
 ("Helezon güç sabiti","75 · 1,36","1 BG = 75 kgm/s ; 1,36 BG = 1 kW dönüşümü / metric-HP→kW."),
 ("Bond","E = 10·Wi·(1/√P₈₀ − 1/√F₈₀)","Bilyalı değirmen özgül enerjisi / ball-mill specific energy (kWh/t)."),
 ("lb·ft","× 0,73756","Nm → lb·ft."),
 ("","",""),
 ("SERVİS FAKTÖRÜ / SERVICE FACTOR","Sf","Her uygulamada düzenlenebilir (varsayılan sütunda). / Editable per application (default column)."),
 ("Gerekli nominal moment","Mₙ = Mᴛ · Sf","Katalog nominal momenti bu değerden büyük olmalı. / Catalogue nominal moment must exceed this."),
 ("Motor gücü / Motor power","Pₘ = Pₒ / η","Bir üst standart IEC güce yuvarlanır. / Rounded up to next IEC size."),
 ("IEC standart kW","0,12 · 0,18 · 0,25 · 0,37 · 0,55 · 0,75 · 1,1 · 1,5 · 2,2 · 3 · 4 · 5,5 · 7,5 · 11 · 15 · 18,5 · 22 · 30 · 37 · 45 · 55 · 75 · 90 · 110 · 132 · 160 · 200 · 250 · 315 · 355","Standart motor güçleri / standard motor sizes."),
]
for it,val,exp in rows:
    r=put(ws4,[it,val,exp],zebra=False)
    if it.isupper() and it:
        for c in range(1,4): ws4.cell(row=r,column=c).font=bold
for i,w in enumerate([30,30,70],1): ws4.column_dimensions[get_column_letter(i)].width=w
ws4.freeze_panes="A3"

# ---- Sheet 5: Read me ----
ws5=wb.create_sheet("Oku-Read me")
banner(ws5,"Nasıl kullanılır / How to use",1)
notes=[
 "Bu çalışma kitabı, İ-MAK Redüktör Seçici uygulamasındaki (telefon + Windows) tüm formülleri belgeler.",
 "This workbook documents every formula in the I-MAK Gearbox Selector app (phone + Windows).",
 "",
 "Her uygulamanın KENDİ sayfası, KENDİ şeması ve KENDİ formülü vardır (İ-MAK Ar-Ge Teknik Eğitim Sunumu esas alınmıştır).",
 "Each application has its OWN page, OWN diagram and OWN formula (based on the I-MAK R&D technical training deck).",
 "",
 "Sayfalar / Sheets:",
 "  • Uygulamalar / Applications — her uygulama, formülü (TR/EN), Sf, η, önerilen ürün ve örnek sonuç.",
 "  • Değişkenler / Variables — her giriş alanı ve varsayılan değeri (yeniden ayarlamak için burayı düzenleyin).",
 "  • Örnek Hesap / Worked example — varsayılan değerlerle adım adım hesap (uygulamadaki 'Hesap adımları' ile aynı).",
 "  • Sabitler / Constants — g, 9550, 367, Bond, servis faktörü ve IEC motor güçleri.",
 "",
 "Hesap akışı / How it computes:",
 "  1. Formül çıkış momenti Mᴛ (Nm) veya çıkış gücü Pₒ (kW) ile çıkış devri n'yi verir.",
 "     The formula returns output moment Mᴛ (Nm) or output power Pₒ (kW) and output speed n.",
 "  2. Mᴛ = 9550·Pₒ/n   veya / or   Pₒ = Mᴛ·n/9550.",
 "  3. Motor gücü Pₘ = Pₒ/η → bir üst IEC güç. / Motor power = Pₒ/η → next IEC size.",
 "  4. Gerekli nominal moment Mₙ = Mᴛ·Sf. / Required nominal moment = Mᴛ·Sf.",
 "",
 "Bir formülü değiştirmek isterseniz: uygulamayı ve yeni denklemi/varsayılanları bana bildirin ya da",
 "'Değişkenler' sayfasını düzenleyip geri gönderin; her iki uygulama sürümünü buna göre yeniden üretirim.",
 "To change a formula: tell me the application and the new equation/defaults, or edit the 'Variables' sheet",
 "and send it back — I will regenerate both app versions to match.",
 "",
 "NOT / NOTE: Süreç sonuçları saha tahminidir; nihai seçimi İ-MAK kataloğu ile doğrulayın.",
 "Process results are field estimates; confirm the final choice against the I-MAK catalogue.",
]
for tx in notes:
    ws5.append([tx]); c=ws5.cell(row=ws5.max_row,column=1); c.alignment=Alignment(wrap_text=True,vertical="top")
    if tx.endswith(":") or tx.startswith("NOT") or tx.startswith("Hesap") : c.font=bold
ws5.column_dimensions["A"].width=115

wb.save(OUT)
print("saved",OUT,"| sheets:",wb.sheetnames)
