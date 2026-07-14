#!/usr/bin/env python3
"""Merge all enriched decision-makers from the batch workbooks into one master contact list."""
import csv, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = [
    ("Denmark","Food",       "/home/user/zak/market-research/I-MAK_Denmark_Batch1_Food.xlsx"),
    ("Denmark","Agriculture","/home/user/zak/market-research/I-MAK_Denmark_Batch2_Agriculture.xlsx"),
    ("South Africa","Food",  "/home/user/zak/market-research/south-africa/I-MAK_SouthAfrica_Batch1_Food.xlsx"),
]

def read_dm(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Decision Makers"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        company = row[0]; full_name = row[2] if len(row) > 2 else None; title = row[3] if len(row) > 3 else None
        # real contact rows have Company + Full Name + Title; note/footer rows don't
        if not company or not full_name or not title:
            continue
        email = row[4] if len(row) > 4 else ""
        status = row[5] if len(row) > 5 else ""
        linkedin = row[6] if len(row) > 6 else ""
        location = row[7] if len(row) > 7 else ""
        out.append([company, full_name, row[1], title, email or "", status or "", linkedin or "", location or ""])
    wb.close()
    return out

master = []
for country, sector, path in SRC:
    for rec in read_dm(path):
        master.append([country, sector] + rec)

# sort: country, company, then verified emails first
status_rank = {"verified":0,"extrapolated":1,"unavailable":2,"":3}
master.sort(key=lambda r: (r[0], r[2].lower(), status_rank.get(r[7],3), r[3]))

HEAD = ["Country","Sector","Company","Full Name","Role","Title","Business Email","Email Status","LinkedIn URL","Location"]

# CSV
csv_path = "/home/user/zak/market-research/I-MAK_Master_Contacts.csv"
with open(csv_path,"w",newline="",encoding="utf-8") as f:
    w = csv.writer(f); w.writerow(HEAD); w.writerows(master)

# Excel
NAVY="1F3864"; GREEN="C6EFCE"; YEL="FFEB9C"; RED="FFC7CE"; LGREY="F2F2F2"; BLUE="2E5496"
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(wrap_text=True,vertical="top"); center=Alignment(horizontal="center",vertical="center")
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Master Contacts"
for c,h in enumerate(HEAD,1):
    cell=ws.cell(row=1,column=c,value=h); cell.font=Font(bold=True,color="FFFFFF",size=11)
    cell.fill=PatternFill("solid",fgColor=NAVY); cell.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center"); cell.border=border
r=2
for rec in master:
    for c,val in enumerate(rec,1):
        cell=ws.cell(row=r,column=c,value=val); cell.alignment=wrap; cell.border=border
        if r%2==0: cell.fill=PatternFill("solid",fgColor=LGREY)
    stc=ws.cell(row=r,column=8); st=rec[7]
    stc.fill=PatternFill("solid",fgColor=GREEN if st=="verified" else YEL if st=="extrapolated" else RED)
    r+=1
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(HEAD))}{len(master)+1}"
for i,w in enumerate([12,12,34,24,18,40,36,13,44,18],1):
    ws.column_dimensions[get_column_letter(i)].width=w

# summary sheet
ws2=wb.create_sheet("Summary")
verified=sum(1 for m in master if m[7]=="verified")
by_country={}
for m in master: by_country[m[0]]=by_country.get(m[0],0)+1
lines=[
 ("I-MAK Reduktor — Master Contact List",True),
 (f"Total contacts: {len(master)}",False),
 (f"Verified business emails: {verified}",False),
 (f"With email (any status): {sum(1 for m in master if m[6])}",False),
 ("",False),
 ("By country:",True),
]+[(f"  {k}: {v}",False) for k,v in sorted(by_country.items())]+[
 ("",False),
 ("Email status legend: verified = deliverable; extrapolated = pattern-inferred; unavailable = no email (LinkedIn provided).",False),
 ("Source: Apollo People Match enrichment across batch workbooks. Sorted by country, company, verified-first.",False),
]
for i,(t,bold) in enumerate(lines,1):
    c=ws2.cell(row=i,column=1,value=t); c.font=Font(bold=bold,size=13 if (bold and i==1) else 11,color=BLUE if bold else "000000")
ws2.column_dimensions['A'].width=100

out="/home/user/zak/market-research/I-MAK_Master_Contacts.xlsx"
wb.save(out)
print("Saved",out,"and",csv_path,"| total contacts:",len(master),"| verified:",verified)
