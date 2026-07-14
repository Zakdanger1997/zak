#!/usr/bin/env python3
"""Build the South Africa Batch 1 (food) Excel workbook for I-MAK Reduktor."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUP="SEW-Eurodrive / NORD / Bonfiglioli / Bauer (INFERRED)"
# Company, Website, HQ, Industry, Machinery, GearboxApp, Revenue, Employees, Export, Evidence, Score, Sources
rows=[
["Filmatic Packaging Systems","filmatic.com","Paarl, W. Cape","Liquid filling / bottling lines","Rotary fillers, cappers, labellers, conveyors, shrink wrap, palletisers; turnkey 800-36,000 BPH","Filler/capper carousels, conveyor drives, palletiser axes - geared motors throughout","~USD 26M (INFERRED)","~120 (INFERRED)","28 African countries + global","Anchor SA filling-line OEM since 1979; Trepko Group (DK) subsidiary, manufactures in Paarl",9,"filmatic.com; dnb.com"],
["Acepak","acepak.com","Cape Town (Montague Gardens)","End-of-line packaging","Case packers, shrink-wrap, palletisers, banders, pouch fillers","Palletiser lift/rotate axes, conveyor & case-erector drives","","~30-80 (INFERRED)","36+ countries","Family-owned 40+ yrs; designs/assembles/tests all equipment in Cape Town",9,"acepak.com; linkedin"],
["Macadams International","macadams.co.za","Cape Town (Blackheath)","Bakery equipment OEM","Ovens, spiral/dough mixers, dividers, moulders, provers, sheeters, slicers","Mixer drives, prover/oven conveyor bands, divider/moulder/sheeter reductions - high content","","Large (17,500 m2 plant)","65+ countries","Est. 1904; largest SA bakery-equipment OEM; single best food prospect",9,"macadams.co.za; brabys"],
["JF Equipment","jfequipment.com","Roodepoort, JHB","Poultry & red-meat processing","Abattoir lines, pluckers, evisceration, chillers, conveyors, monorails, packing","Conveyor/monorail drives, plucker rotors, chiller augers","<USD 5M (LOW)","11-50 (LOW)","~15 countries","Est. 1991; SA's leading stainless processing manufacturer, 365 abattoirs built",8,"jfequipment.com; D&B"],
["Modern Processing Equipment (MPE)","poultryprocessing.co.za","Bellville, Cape Town","Poultry/red meat/fruit/dairy","Overhead conveyors/monorails, drum pluckers, scald tanks, evisceration, stunners","Overhead conveyor drives, plucker drum drives, packing-line drives","","","Africa (INFERRED)","Founded 1997; manufactures monorail/conveyor systems + rotating pluckers",8,"poultryprocessing.co.za; FB"],
["STEGEL Engineering","stegel.co.za","Johannesburg","Dairy/beverage/pharma processing","Stainless mixing/process tanks, top/side-entry mixers, high-shear mixers, homogenisers","Agitator/mixer gearboxes, homogeniser & high-shear drives","","","","Builds agitated/high-shear vessels + homogenisers - each needs mixer drive",8,"stegel.co.za"],
["Anderson Engineering","andersoneng.co.za","Pietermaritzburg, KZN","Dairy/food/beverage/pharma","Cheese vats, milk-cooling/mixing tanks, agitated vessels, magnetic mixers, tankers","Agitator/mixer drives, cheese-vat stirrer drives","~USD 17.8M (LOW)","14 (LOW)","France, Kenya, Nigeria, DRC","Est. 1958, ISO 9001/3834; dairy specialist building agitated vessels",8,"andersoneng.co.za; crown.co.za"],
["Roff Milling (Roff Industries)","roff.co.za","Kroonstad, Free State","Maize milling machinery OEM","Compact/turnkey maize mills, roller mills, hammer mills, crushers, malt mills","Roller-mill drives, conveyors, elevators, sifters, feeders - geared motors throughout","~USD 6.4M (INFERRED)","51-200 (INFERRED)","Sub-Saharan Africa + E. Europe","Est. 1991; turnkey mills exported across 12+ African countries",8,"roff.co.za; linkedin"],
["ABC Hansen Africa","abchansenafrica.co.za","Pretoria, Gauteng","Grain/maize milling & handling","Roller mills, Hippo hammer mills, stone mills, bucket elevators, conveyors, sifters, silos","Roller-mill drives, plansifter eccentrics, conveyor & elevator geared motors, augers","","~5,000 m2 (INFERRED)","Many African & overseas","Since 1991; built 200+ grain mills; makes Hippo mills (since 1928)",8,"abchansenafrica.co.za"],
["Facet Engineering","facetengineering.co.za","Krugersdorp, Gauteng","Bulk material handling + food conveying/blending","Bucket elevators, belt/chain/screw/vibratory conveyors, ribbon blenders, grain cleaners","Elevator head-shaft & conveyor geared motors, blender shafts - core driver","","~40-80 (INFERRED)","Africa","Est. 1989; one of the largest ranges of conveyors in SA; serves grain/milling/food",8,"facetengineering.co.za"],
["Jones Industrial Mixers","jonesmixers.co.za","Edenvale, JHB","Mixing/blending machinery","Ribbon blenders, Z-blade mixers, double-cone blenders, IBC mixers, agitators, dispersers","Geared motors/reducers drive agitator & ribbon/paddle shafts and rotating vessels","","~20-50 (INFERRED)","Southern Africa","7 decades; local mixing/blending OEM; food among industries",8,"jonesmixers.co.za; engnet"],
["CFAM Technologies","cfaminternational.com","Potchefstroom, NW","Food/feed extrusion machinery","Twin-screw extruders, complete food/feed/pet-food/snack extrusion plants","Heavy-duty reduction gearboxes drive twin screws; geared motors on augers/conveyors","R11-100M (INFERRED)","51-200 (INFERRED)","18 countries (UK, NA, Asia, EU, Africa)","NWU spin-off 2007; clients PepsiCo, Tiger Brands, Premier",8,"cfaminternational.com; sagrainmag"],
["National Packaging Systems (NPS)","nationalpackagingsystems.co.za","Durban (Westmead), KZN","VFFS / fillers","VFFS, volumetric & auger fillers, stick-pack, flowwrappers, screw feeders, elevators","Gearmotors on VFFS film-draw & sealing jaws, auger fillers, elevators, conveyors","","~20-50 (INFERRED)","Primarily SA","Est. 1983, raw material to finished machine under one roof; clients Unilever, Tiger Brands",8,"nationalpackagingsystems.co.za"],
["Powercon Handling Systems","powercon.co.za","Durban (Queensburgh), KZN","Conveying / material handling","Belt/roller/chain/monorail/steel-belt conveyors, bag stackers, truck loaders","Geared motors are core drive on every powered conveyor, stacker & elevator","","~30-70 (INFERRED)","Worldwide export capability","Since 1982; designs & manufactures conveyors in Durban; high gearbox intensity",8,"powercon.co.za"],
["Egli Precision Engineering","egli.co.za","Pietermaritzburg, KZN","VFFS / baling / snack lines","VFFS machines, baling machinery, complete maize-snack manufacturing lines","Gearmotors on VFFS drives, baler mechanisms, snack-line conveyors/extruder feeds","","~30-80 (INFERRED)","Exporter (Made in Africa)","Founded 1973; one of SA's largest OEM VFFS manufacturers",8,"egli.co.za; ivote"],
["Schullpak CC","schullpak.co.za","Cape Town (+ JHB)","Bottling-line automation & conveyors","Bottle/can/crate/pallet conveyors (up to 72,000 bph), star-wheels, scrolls, feed","Conveyor drive units, star-wheel & timing-screw drives","","SME","Southern Africa (INFERRED)","SA manufacturer 20+ yrs; serves wineries, craft breweries, mass bottling",8,"schullpak.co.za"],
["Gossamer Packaging Machinery","gossamer.co.za","Somerset West, W. Cape","Packhouse end-of-line automation","Robot palletisers, tray/carton erectors, bin handlers, case sealers, conveyors, gantries","Palletiser & conveyor/gantry drives, carton-erector mechanisms","","~30-80 (INFERRED)","SA + USA/Australia installs","30+ yrs; locally manufactured in SA; robot palletiser at Two-a-Day Grabouw",8,"gossamer.co.za; bizcommunity"],
["Maize Master","maizemaster.co.za","Kroonstad, Free State","Maize/grain milling machinery","Roller mills (500 kg/h-3.5 t/h), turnkey & containerised mills, feed-mill equipment","Roller-mill & feed drives, conveyors, elevators, mixers","","SME (INFERRED)","Across Africa","Est. 1988; claims world's first small commercial roller mill; turnkey mills",7,"maizemaster.co.za"],
["PACCO (P.A. Cuthbert & Co)","pacuthbert.co.za","Modderfontein, JHB","Powder mixing/blending machinery","Ribbon blenders, V-blenders, double-cone blenders, mills, fluid-bed dryers","Geared drives rotate blender agitator shafts and tumble-blender vessels","","~30-70 (INFERRED)","Africa (INFERRED)","In-house SS fabrication; GMP food & pharma; local manufacturer",7,"pacuthbert.co.za"],
["Snack Quip","snackquip.com","Benoni, JHB","Snack-food line machinery","Extrusion lines, pellet frying, popcorn/cornflake lines, twin-screw extruder, fryers, baggers","Geared motors on extruder drives, conveyors, augers, fryer paddles","","~10-30 (INFERRED)","S. Africa, Malawi, Kenya","Supplies Simba (Frito-Lay) & Willards; 40+ yrs combined experience",7,"snackquip.com; howwemadeitinafrica"],
["Arcrite Engineering","arcrite.co.za","Cape Town","Blending & conveying machinery","Ribbon blenders (100-3000 L), screw conveyors, bucket elevators, materials handling","Geared reducers drive ribbon-blender shafts & screw-conveyor drives; VSD options","","~30-60 (INFERRED)","Africa (INFERRED)","40+ yrs (est. 1984); food, spice, feed, cake-mix blending; BBBEE L1",7,"arcrite.co.za"],
["Mixtec","mixtec.com","Germiston (Meadowdale)","Industrial agitators/mixers","Top/side/bottom-entry agitators & mixers, impellers, shafts, drive assemblies","OPENLY integrates 3rd-party gearbox makers' units as mixer drive heads","","~50-150 grp (INFERRED)","USA, AU, UK, Malaysia, Chile, NZ (7 plants)","SA-HQ mixer OEM openly multi-sourcing gearboxes = ideal target",7,"mixtec.com"],
["Afromix (AFX Holdings)","afromix.co.za","Boksburg, Gauteng","Industrial agitators/mixers","Agitators/mixers 0.25-750 kW, peristaltic pumps, impellers, shafts","Large reduction gearboxes on high-torque agitator drives across full range","","~30-80 (INFERRED)","Europe, NA, Australia, Africa","5,000+ installations; food & beverage, pharma, chemical",7,"afromix.co.za; afxholdings"],
["Calculus-TAC","calculus-tac.co.za","Cape Town","Weighing / bagging / filling","Bagging (grain/sugar/fertiliser), bulk-bag & carousel fillers, multihead weighers, belt scales","Gearmotors on bag conveyors, carousel indexing, bulk-fill augers, belt-scale drives","","~20-50 (INFERRED)","SA + regional","Manufacturing Calculus range ~41 yrs (since 1981)",7,"calculus-tac.co.za"],
["PMD Packaging","pmdpackaging.co.za","Randburg, JHB","VFFS / bagmakers / filling","Automated bagmakers (pillow/doypack/stick-pack), Sigma VFFS, powder/granular fillers","Gearmotors on VFFS film feed & jaws, filler augers, conveyors","","~15-40 (INFERRED)","SA base","Self-described SA manufacturer of bagmakers + fillers (Sigma range)",7,"pmdpackaging.co.za"],
["Autopac Machinery","autopac.co.za","Johannesburg","Form-fill-seal / snack lines","Draw-down & belt FFS, ice-lolly machines, peanut/snack lines, ancillaries","Gearmotors on FFS film/belt drives, snack-line conveyors & augers","","~15-40 (INFERRED)","25+ countries","Est. 1986; in-house fabrication/assembly/commissioning",7,"autopac.co.za"],
["Inpakt Group","inpaktgroup.co.za","Durban, KZN","Filling, packaging, palletizing, intralogistics","Filling & packaging machines, complete lines, palletizing, process systems","Palletizer & conveyor/intralogistics drives, filling carousels","","11-50 (INFERRED)","Sub-Saharan Africa","Designs, develops and manufactures machines and complete lines, est. 2003",7,"inpaktgroup.co.za"],
["Innovative Processing Solutions (IPS)","ips-sa.co.za","Germiston, JHB","Agitators/mixing machinery","Modular agitators (top/side/bottom), tank/eductor mixers, powder mixers","Geared drive heads on agitator/mixer shafts","","~10-40 (INFERRED)","Africa (INFERRED)","SA manufacturer serving pharma, food & beverage mixing",6,"ips-sa.co.za"],
["MaxMix","maxmix.co.za","Cape Town (Blackheath)","Mixing/dispersing machinery","Ribbon blenders, high-shear & rotor/stator mixers, dispersers, lab mixers","Geared motors drive blender shafts; direct/geared on dispersers","","~20-50 (INFERRED)","Exports to Africa","Founded 1996; 3,500 m2 plant; food & beverage among industries",6,"maxmix.co.za"],
["GVTEC (Grotto + Velo)","gvtec.co.za","Wellington, W. Cape","Stainless process tanks/vessels","Tanks 100 L-400,000 L; fermentation/dairy/wine vessels (some agitated)","Tank agitators/mixers, must pumps - mixer gearboxes","","~90 (CONFIRMED)","Five continents","Large SA fabricator (Grotto 50 yrs + Italian Velo); agitated vessels",6,"gvtec.co.za"],
["Albrecht Machinery","albrechtmachinery.co.za","Paarl, W. Cape","Food/poultry freezing & handling","Locally-built spiral freezers (IQF chicken), carton freezers, conveyor belts","Spiral-freezer belt/drum drives, conveyor drives","","","Regional (INFERRED)","Est. 1949; builds spiral freezers & conveyors in Paarl",7,"albrechtmachinery.co.za"],
["Basils Business Opportunities (BBO)","bbosa.co.za","Germiston (Elsburg)","Poultry abattoir + agro/food processing","Pluckers, shackle/conveyor lines, fruit finishers/pulpers, peanut-butter & oil-seed lines, roasters","Geared motors on pluckers, conveyors, augers, mixing-tank agitators","","~10-30 (INFERRED)","10+ African countries","SS agro-processing OEM to SABS standard",6,"bbosa.co.za; agri4africa"],
["Drotsky","drotsky.co.za","Alberton, Gauteng","Hammer mills & feed mixers","Hammer mills (M/S/T), maize/feed mixers (ribbon/paddle)","Feed-mixer paddle/ribbon drives, auger & conveyor geared motors (hammer mills belt-driven)","","SME (INFERRED)","~15 African countries + NZ/AU/NL","Est. 1962; 60,000+ hammer mills sold; still manufacturing in Alberton",6,"drotsky.co.za"],
["Tombake","tombake.co.za","Johannesburg (+ CT, Durban)","Bakery equipment (part mfr / part importer)","Dough & cake mixers, dividers, provers, ovens (own factory + imports)","Mixer drives, prover/oven conveyors, divider drives","","Mid (INFERRED)","S. Africa & Africa","Manufactures bakery equipment from scratch + imports; supplies retail groups",6,"tombake.co.za; tabj"],
["Micro Brewing SA","microbrewingsa.co.za","KZN (INFERRED)","Brewery / brewhouse equipment","Brewhouse vessels, mash tuns, turnkey micro-breweries & brewpubs","Mash-tun rakes/agitators, lauter rakes","","SME","S. Africa (INFERRED)","Brewery-equipment manufacturer; second-source verification thin - flag before outreach",6,"microbrewingsa; businesstech"],
["NCC Siyaphambili","ncccomp.com","Kempton Park, Gauteng","Conveyor systems & components (beverage/canning)","Conveyor systems + engineered plastic components (star-wheels, guides) for bottling/canning","Conveyor drive units on assembled systems","","SME","Southern Africa (INFERRED)","SA manufacturer since 1997; brewing/beverage/canning focus",6,"ncccomp.com; b2bhint"],
["MA Engineering","maengineer.co.za","Alberton (Alrode), JHB","Filling / capping / labelling","Piston fillers, cap-twist machines, labellers, line automation","Gearmotors on capping spindles, labeller drives, conveyor/indexing","","~5-20 (INFERRED)","SA-focused","All products designed & manufactured within SA, locally sourced parts",6,"maengineer.co.za"],
["Mechaneer","mechaneer.co.za","South Africa","Robotic packaging / conveying","Automated bag-placing (5-80 kg), robotic palletising, motor-driven conveyors","WORM-GEAR AC motors documented on conveyors; gearmotors on bag-place/palletising","","~5-20 (INFERRED)","SA (INFERRED)","Site documents worm-gear drives - rare confirmed gearbox signal",6,"mechaneer.co.za"],
["Dairy Dynamics","dairydynamics.co.za","KZN (INFERRED)","Small-scale dairy processing","Milk process/packaging plants from 50 L, cream separators, homogenising pumps, fillers","Separator bowl drives, filler carousel drives, pump drives","","","Southern Africa; 2,000+ units","Since ~1996; builds/assembles separators, fillers, plants (partly imports)",6,"dairydynamics.co.za; machinio"],
["Saturn Stainless Industries","saturnstainless.co.za","Paarl (+ Benoni)","Stainless tanks & equipment (wine)","Wine/beverage stainless tanks, custom vessels, process equipment","Agitated tanks / mixing equipment - geared mixer drives","","1,500 m2 facility","Local + international","SA manufacturer since 1979; gearbox use where agitators fitted",5,"saturnstainless.co.za"],
["Millcon Engineering","millconengineering.co.za","Gauteng (INFERRED)","Milling machinery components & turnkey plants","Spares/components for elevators, conveyors, milling; turnkey plant design","Supplies geared-motor-driven elevators/conveyors within turnkey mills","","Small (INFERRED)","SA / Africa (INFERRED)","Engineering supply for milling; more consultant/integrator than pure OEM",5,"millconengineering.co.za"],
["Southern Bakery Equipment","southernbakery.co.za","Johannesburg (Westgate)","Bakery equipment (new + reconditioned)","Ovens, provers, mixers, moulders, dividers","Mixer/prover/divider geared drives (assembler/reconditioner)","","Small (INFERRED)","Mainly SA (INFERRED)","Est. 1996 by ex-W&P engineers; limited own-manufacture",4,"southernbakery.co.za"],
]

HEADERS=["#","Company Name","Website","Headquarters","Industry","Machinery Produced","Gearbox Application","Est. Annual Revenue","Employees","Export Markets","Existing Gearbox Suppliers (inferred)","Evidence","Buying Potential (1-10)","Source Links"]
NAVY="1F3864"; BLUE="2E5496"; LGREY="F2F2F2"; GREEN="C6EFCE"; YEL="FFEB9C"; RED="FFC7CE"
hdr_font=Font(bold=True,color="FFFFFF",size=11); hdr_fill=PatternFill("solid",fgColor=NAVY)
title_font=Font(bold=True,color="FFFFFF",size=16)
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(wrap_text=True,vertical="top"); center=Alignment(horizontal="center",vertical="center")
def score_fill(s): return PatternFill("solid",fgColor=GREEN if s>=8 else YEL if s>=6 else RED)
def style_header(ws,ncols,row=1):
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c); cell.font=hdr_font; cell.fill=hdr_fill
        cell.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center"); cell.border=border
def write_table(ws,headers,data,start_row=1,score_col=None):
    for c,h in enumerate(headers,1): ws.cell(row=start_row,column=c,value=h)
    style_header(ws,len(headers),start_row); r=start_row+1
    for rec in data:
        for c,val in enumerate(rec,1):
            cell=ws.cell(row=r,column=c,value=val); cell.alignment=wrap; cell.border=border
            if r%2==0: cell.fill=PatternFill("solid",fgColor=LGREY)
        if score_col:
            sc=ws.cell(row=r,column=score_col); sc.fill=score_fill(rec[score_col-1]); sc.alignment=center; sc.font=Font(bold=True)
        r+=1
    ws.freeze_panes=ws.cell(row=start_row+1,column=1)

wb=openpyxl.Workbook()
ws=wb.active; ws.title="Read Me"
ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=110
ws.merge_cells('A1:B1'); ws['A1']="I-MAK Reduktor  |  South Africa Machinery OEM Prospecting  -  BATCH 1: Food"
ws['A1'].font=title_font; ws['A1'].fill=PatternFill("solid",fgColor=NAVY); ws['A1'].alignment=center; ws.row_dimensions[1].height=30
readme=["","SCOPE","SA Batch 1 of 6 - food sector: dairy/meat, beverage/winery/bottling, bakery & milling,",
"general food processing/mixing, and packaging. 42 unique South African OEMs.","",
"TABS","  Company Database - all 42 companies","  Top Prospects - ranked by score",
"  By Industry / By Region - groupings","  Gearbox Suppliers - incumbents to displace","",
"KEY MARKET INSIGHT","  SEW-Eurodrive DOMINATES the SA installed base (local assembly Nelspruit + Johannesburg).",
"  I-MAK's angle = price/lead-time on standard helical, worm & helical-bevel gearmotors vs SEW.",
"  Also present: NORD, Bonfiglioli, Bauer, SM-Cyclo, Flender.","",
"METHOD & CONFIDENCE","  2+ source verification; CONFIRMED vs INFERRED; blanks = unknown, not fabricated.",
"  SA food-machinery is a thinner OEM space than Denmark - many importers/agents were excluded.",
"  Firmographics largely undisclosed for private SMEs (mostly INFERRED) - verify via CIPC before qualification.",
"  Score: green 8-10 (priority), yellow 6-7, red <=5.","",
"CROSS-LINK: Filmatic (Paarl) is owned by Denmark's Trepko Group - potential relationship lever.",
"NOTE: SA's biggest OEM sectors come later - Batch 3 (Mining) and Batch 4 (Materials Handling).",
"DECISION-MAKERS: Apollo enrichment available on request (SA coverage thinner than DK)."]
for i,line in enumerate(readme,start=2):
    c=ws.cell(row=i,column=2,value=line)
    c.font=Font(bold=True,size=12,color=BLUE) if (line.isupper() and line.strip()) else Font(size=11)

ws=wb.create_sheet("Company Database")
data=[[i+1]+rows[i][:9]+[SUP]+rows[i][9:] for i in range(len(rows))]
write_table(ws,HEADERS,data,score_col=13)
for i,w in enumerate([4,30,24,22,26,34,38,20,16,22,28,40,12,24],1):
    ws.column_dimensions[get_column_letter(i)].width=w

ws=wb.create_sheet("Top Prospects")
order=sorted(range(len(rows)),key=lambda i:-rows[i][10])
tp=[[rank,rows[i][0],rows[i][3],rows[i][2],rows[i][10],rows[i][5],rows[i][9]] for rank,i in enumerate(order,1)]
write_table(ws,["Rank","Company","Industry","HQ","Score","Gearbox Application","Evidence"],tp,score_col=5)
for i,w in enumerate([6,32,28,24,8,42,46],1): ws.column_dimensions[get_column_letter(i)].width=w

ws=wb.create_sheet("By Industry")
industry_map={
 "Bakery & Milling":[2,7,8,9,17,32,33,41],
 "Packaging / Filling / Conveying":[0,1,12,13,14,15,16,23,24,25,26,36,37],
 "Mixing / Agitators / Blending":[10,18,20,21,22,27,28],
 "Dairy / Meat / Beverage processing":[3,4,5,6,29,30,31,34,35,38,39],
 "Food extrusion / snack":[11,19],
}
r=1
for ind,idxs in industry_map.items():
    ws.cell(row=r,column=1,value=ind).font=Font(bold=True,size=13,color="FFFFFF")
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    for c in range(1,5): ws.cell(row=r,column=c).fill=PatternFill("solid",fgColor=BLUE)
    r+=1
    for c,h in enumerate(["Company","HQ","Machinery Produced","Score"],1): ws.cell(row=r,column=c,value=h)
    style_header(ws,4,r); r+=1
    for i in sorted(set(idxs),key=lambda i:-rows[i][10]):
        rec=[rows[i][0],rows[i][2],rows[i][4],rows[i][10]]
        for c,val in enumerate(rec,1):
            cell=ws.cell(row=r,column=c,value=val); cell.alignment=wrap; cell.border=border
        sc=ws.cell(row=r,column=4); sc.fill=score_fill(rec[3]); sc.alignment=center; sc.font=Font(bold=True); r+=1
    r+=1
for i,w in enumerate([32,24,50,8],1): ws.column_dimensions[get_column_letter(i)].width=w

ws=wb.create_sheet("By Region")
def region_of(hq):
    h=hq.lower()
    if any(k in h for k in ["cape","paarl","wellington","somerset","stellenbosch","w. cape","blackheath","montague"]): return "Western Cape"
    if any(k in h for k in ["durban","kzn","pietermaritzburg","westmead","queensburgh"]): return "KwaZulu-Natal"
    if any(k in h for k in ["kroonstad","free state"]): return "Free State"
    return "Gauteng / Inland"
from collections import defaultdict
reg=defaultdict(list)
for i,r0 in enumerate(rows): reg[region_of(r0[2])].append(i)
rr=1
for region in ["Gauteng / Inland","Western Cape","KwaZulu-Natal","Free State"]:
    if region not in reg: continue
    ws.cell(row=rr,column=1,value=f"{region}  ({len(reg[region])} companies)").font=Font(bold=True,size=13,color="FFFFFF")
    ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=4)
    for c in range(1,5): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=BLUE)
    rr+=1
    for c,h in enumerate(["Company","HQ","Industry","Score"],1): ws.cell(row=rr,column=c,value=h)
    style_header(ws,4,rr); rr+=1
    for i in sorted(reg[region],key=lambda i:-rows[i][10]):
        rec=[rows[i][0],rows[i][2],rows[i][3],rows[i][10]]
        for c,val in enumerate(rec,1):
            cell=ws.cell(row=rr,column=c,value=val); cell.alignment=wrap; cell.border=border
        sc=ws.cell(row=rr,column=4); sc.fill=score_fill(rec[3]); sc.alignment=center; sc.font=Font(bold=True); rr+=1
    rr+=1
for i,w in enumerate([32,24,30,8],1): ws.column_dimensions[get_column_letter(i)].width=w

ws=wb.create_sheet("Gearbox Suppliers")
gs=[
["SEW-Eurodrive","Germany (SA assembly)","DOMINANT SA incumbent - local assembly Nelspruit + JHB; default drive on most food OEMs","Price + lead-time on standard helical/worm/bevel"],
["NORD Drivesystems","Germany","Conveyors, VFFS, packaging lines, milling handling","Hygienic/stainless, delivery"],
["Bonfiglioli","Italy","Mixers, extruders, heavy conveying","Cost on high-torque/extruder units"],
["Bauer (Danfoss)","Germany","Compact gearmotors on fillers, conveyors, smaller machines","Local relationship + custom"],
["SM-Cyclo (Sumitomo)","Japan","Bagging/indexing, some conveying","Cyclo ratio niche"],
["Flender (Siemens)","Germany","Large extruder/mill drives","Hard at high power; target site engineering"],
]
write_table(ws,["Incumbent Supplier","Origin","Typical position in SA food machinery","I-MAK displacement angle"],gs)
ws.cell(row=len(gs)+3,column=1,value="All attributions INFERRED (no public BOM) except Mechaneer (documents worm-gear) & Mixtec (openly multi-sources). SEW is the primary account incumbent.").font=Font(italic=True,color="C00000")
for i,w in enumerate([22,20,54,40],1): ws.column_dimensions[get_column_letter(i)].width=w

out="/home/user/zak/market-research/south-africa/I-MAK_SouthAfrica_Batch1_Food.xlsx"
wb.save(out)
print("Saved",out,"| companies:",len(rows),"| tabs:",wb.sheetnames)
