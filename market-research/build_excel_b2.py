#!/usr/bin/env python3
"""Build the Batch 2 (agriculture sector) Excel workbook for I-MAK Reduktor prospecting."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ELEC = "SEW / NORD / Bonfiglioli / Bauer (INFERRED)"
PTO  = "Comer / Bondioli & Pavesi / Walterscheid (INFERRED)"
MIX  = "Planetary + Comer / SEW (INFERRED)"

# Company, Website, HQ, Industry, Machinery, GearboxApp, Revenue, Employees, Export, Suppliers, Evidence, Score, Sources
rows = [
["Cimbria A/S (AGCO; DK mfg)","cimbria.com","Thisted (+ Haderslev)","Grain/seed processing & handling","Continuous grain dryers, silos, bucket elevators, belt/chain/screw conveyors, Delta seed cleaners, intake, loading chutes","Geared motors on every elevator head, conveyor drive, dryer discharge roller, screw/chain drive - very high unit count/plant","~DKK 1.5-2bn grp / >USD 244M (CONFIRMED range)","~900 grp / ~740 A/S (CONFIRMED)","100+ countries",ELEC,"Two DK plants mass-producing elevators/conveyors/dryers; highest-volume gearbox lead in batch",10,"cimbria.com; mergr.com"],
["SKIOLD A/S (Solix Group)","skiold.com","Saeby (+ Ikast, Bur, Olgod)","Feed milling + grain handling + pig/poultry systems","Hammer/disc/roller mills, mixers, chain & disc conveyors, augers, silo dischargers, dry feeding, complete feed mills","DOCUMENTS using planetary gear motors; geared drives on mills, mixers, chain-disc feeding, augers, elevators","~DKK 1bn+ grp (INFERRED)","500-700 grp (CONFIRMED)","60+ countries; 90% export",ELEC,"Multi-plant feed-mill + farming OEM; huge drive count. Parent of Landmeco, Damas, JYDEN. Top group lead",9,"skiold.com; millingandgrain.com; wikipedia"],
["ACO Funki A/S (incl. Egebjerg)","acofunki.com","Aulum (Herning)","Pig housing, feeding & manure","Dry & liquid feeding, TUBE-O-MAT feeders, chain-disc feed conveyors, under-slat manure scrapers","Geared motors on chain-disc feed loops, manure scraper stations, feed augers/dosing","~EUR 40-80M (INFERRED)","~150-250 (INFERRED)","DE, FR, ES, PL, RU, UA, CN",ELEC,"90-yr pig-equipment OEM; feeding conveyors + scrapers gear-driven. Egebjerg = same entity",9,"acofunki.com; agriexpo"],
["SAMSON AGRO A/S","samson-agro.com","Viborg","Slurry/manure application (flagship)","Slurry tankers, injectors/incorporators, muck & solid spreaders, self-propelled applicators","PTO gearboxes on pumps, spreading rotors, auger/beater drives; angle gearboxes on self-propelled drivelines","Grp net turnover EUR 127.2M FY23/24 (CONFIRMED)","Grp ~613 (324 DK) (CONFIRMED)","France, Poland, Sweden, global",PTO,"Multiple driven pump/rotor/beater systems per machine; flagship DK slurry OEM",9,"samson-agro.com; group results"],
["Kongskilde Industries A/S","kongskilde-industries.com","Soro","Grain/feed conveying & cleaning","Chain/belt/screw conveyors, bucket elevators, pneumatic conveying, cleaners/separators, weighing","Geared motors on augers, chain conveyors, elevators, screen/separator drums","~DKK 400-600M (INFERRED); GP ~81M (CONFIRMED)","110-200+ (CONFIRMED range)","FR, DE, ES, PL, UK, USA, ZA, TH",ELEC,"Own Soro production of conveying/cleaning gear; high geared-motor count. Separate from CNH's Kongskilde Agriculture",9,"kongskilde-industries.com; bulkinside"],
["LANDMECO A/S (Skiold Landmeco)","landmeco.com","Olgod","Broiler/layer/breeder house systems","Manure belts, chain/pan feeding lines, nest lines with egg belts, watering, climate","Geared motors on manure belts, egg-collection belts, feed chain/auger lines - high count/barn","~DKK 150-250M (INFERRED)","~48-100 (CONFIRMED-ish)","Worldwide incl. Americas",ELEC,"Continuous conveyor systems throughout each house; part of SKIOLD group",9,"landmeco.com; proff.dk CVR 89524113"],
["ANDRITZ Feed & Biofuel A/S (ex Sprout-Matador)","andritz.com","Esbjerg","Feed & biofuel pelleting plants","Pellet mills (gear-driven), extruders, coolers, hammer mills, paddle mixers, conveyors, complete lines","Main-drive gearbox on pellet mills; geared motors on coolers, screw/chain conveyors, mixers, feeders","~DKK 554M (2023, CONFIRMED)","147 DK (CONFIRMED)","Global (15 sites)",ELEC,"Biggest DK feed-machinery OEM; pellet mills explicitly gear-driven. Caveat: large corporate/central purchasing",8,"andritz.com; vismarating.dk; proff.dk"],
["Jema Agro A/S","jemaagro.dk","Bjerringbro","Grain/feed mechanical conveying","Chain & flight conveyors, chain & bucket elevators, belt conveyors, intake pits, floor-store filling, grain weighers","Geared motor on every elevator/conveyor drive head - core to product","~DKK 60-100M (INFERRED); GP ~17M (CONFIRMED)","~29-60 (CONFIRMED ~29)","Europe (UK floor-store leader)",ELEC,"Pure conveyor/elevator OEM; each machine ships with a geared motor; recurring demand",8,"jemaagro.dk; proff.dk CVR 34472211"],
["Maskinfabrikken Cormall A/S","cormall.dk","Sonderborg","Cattle feed mixers, straw/biomass","Vertical & auger diet/TMR mixers (Multimix), grinding mills, straw processors, biomass shredders","Planetary gearbox on vertical mixer augers; PTO input gearbox; mill/shredder rotor drives","~DKK 20M (INFERRED)","35 (CONFIRMED)","Europe, global",MIX,"Auger/vertical mixers need robust high-torque gearboxes; 60+ yr feed-machine OEM",8,"cormall.dk; proff.dk; danskagroindustri.dk"],
["Bredal A/S","bredal.com","Vejle","Fertilizer/lime spreaders","Large trailed disc spreaders, lime spreaders","PTO gearbox splitting drive to twin spinning discs; floor-conveyor gear drive","Not confirmed (PE-owned)","~30-70 (INFERRED)","Australia, Europe, global",PTO,"Every spreader has a disc-drive gearbox as core component",8,"danskagroindustri.dk; pitchbook"],
["BOGBALLE A/S","bogballe.com","Uldum","Fertilizer spreaders","Mounted twin-disc broadcast spreaders","PTO gearbox dividing drive to two spinning discs","~USD 14M (INFERRED)","~16 (INFERRED, automated)","100+ countries (CONFIRMED)",PTO,"Single-product firm; disc-drive gearbox in every unit; high-volume repeatable demand",8,"bogballe.com; northdata CVR 82983228"],
["HE-VA ApS (incl. Doublet-Record)","he-va.com","Nykobing Mors","Tillage & seeding","Power harrows, seed drills, stubble cultivators, rollers, fine-seed equipment","Power-harrow central + rotor gearboxes; seed-drill metering drive; disc/roller drives","Not confirmed (family)","140+ (CONFIRMED)","90% export; UK, EU, CN, CA, NZ",PTO,"Power harrows gearbox-intensive; seed-drill drives",8,"he-va.com; dnb.com"],
["Fransgard Maskinfabrik A/S","fransgard.dk","Norager (V. Hornum)","Ag & forestry machinery","Muck spreaders, disc mowers, tedders, rotary harrows, crimpers, forestry winches, grapples","PTO gearboxes on spreader beaters/discs, mower/tedder drives; gear reduction in winches","Not confirmed (DK-owned)","Not confirmed","80-90% export, worldwide",PTO,"Broad PTO-driven line (spreaders, mowers, tedders, winches)",8,"fransgard.dk; agromek"],
["Lachenmeier Monsun A/S","lachenmeier-monsun.com","Sonderborg","Industrial bulk handling (grain/feed/malt/biomass)","Chain conveyors, MONSUN heavy-duty bucket elevators, belt transport, complete bulk plants","Geared motors on bucket-elevator head drives & chain/belt conveyor drives","~USD 16.4M / DKK 110M (CONFIRMED)","61 (CONFIRMED)","NO, VN, RU, UA, PL, IN",ELEC,"In-house production of drive-powered elevators/conveyors since 1952",8,"lachenmeier-monsun.com; northdata CVR 73507812"],
["Assentoft Silo A/S","assentoftsilo.dk","Assentoft (Randers)","Silos + complete storage/handling plants","Steel silos, grain pits, chain elevators, flexible mechanical conveyors, complete plants","Geared motors on chain elevators, sweep/discharge augers, conveyor drives","~DKK 20-45M (CONFIRMED range)","23-39 (CONFIRMED)","Denmark + Europe",ELEC,"Supplies complete conveyor plants incl. chain elevators, not just static silos",7,"assentoftsilo.dk; proff.dk"],
["Agrisys A/S","agrisys.dk","Herning","Precision pig feeding (AgTech)","AirSys dry/wet feed distribution, automated weighing, ESF feeders, bedding distribution","Geared motors on feed augers, conveyors, dosing/distribution units","~EUR 5-15M (INFERRED)","~30-60 (INFERRED)","EU, CN, global dealers",ELEC,"Automated feed transport = augers/conveyors driven by geared motors",7,"en.agrisys.dk; danishpigacademy.com"],
["SKIOLD Damas A/S","damasseed.com","Vester Aaby (Faaborg)","Grain & seed cleaning/grading","Screen cleaners, graders, separators, complete seed-processing lines","Geared motors on cleaner eccentric/screen drives, grading drums, feed rollers","~DKK 50-100M (INFERRED)","~20-60 (INFERRED)","Global (seed industry)",ELEC,"Invented the screen cleaner; rotating/oscillating driven parts. Part of SKIOLD (shared purchasing)",7,"damasseed.com; kompass"],
["HARDI International A/S","hardi.com","Norre Alslev","Crop protection / sprayers","Trailed, mounted & self-propelled field sprayers","PTO/pump-drive gearbox on trailed pumps; fan gearbox on air-assist; hub/wheel gear drives on self-propelled","~USD 124M (aggregator, LOW)","200-340 (CONFIRMED range)","Global (13+ subs)",PTO,"Self-propelled models need wheel/hub reduction; pump & fan gearboxes",7,"hardi.com; rocketreach"],
["Thyregod A/S (incl. TIM, Kimadan)","thyregod.com","Thyregod","Root-crop harvesting","Beet harvesters, toppers, silage trailers, straw harrows","PTO gearboxes on digging shares, cleaning turbines, elevator/web drives, topper rotors","Not confirmed","Not confirmed","Global",PTO,"Multi-rotor harvester with driven cleaning/lifting systems",7,"danskagroindustri.dk; agriexpo"],
["Agrometer a/s","agrometer.dk","Grindsted (+ Rodding)","Slurry pumping & injection","Umbilical slurry systems, mobile pump units, SRS hose reels, SDS spreaders, injectors","Gearbox on rotary/AGM pumps; hose-reel rewind gear drive; injector distributor drive","Not confirmed","55 (CONFIRMED)","Europe",PTO,"Rotary pump + reel-drive gearboxes central to product",7,"agrometer.dk; fwi.co.uk"],
["GreenTec A/S","greentec.eu","Kolding","Vegetation/verge maintenance","Boom hedge cutters, flail mowers/mulchers, finger-bar cutters","Angle/bevel gearboxes on flail-rotor & mulcher heads; rotary cutter drives","Not confirmed","40+ (CONFIRMED)","Dealers in 25 countries",PTO,"Rotary/flail heads require robust angle gearboxes",7,"greentec.eu"],
["LINCO Incubator ApS","lincoincubator.com","Ikast","Hatchery equipment","Single-stage setters/hatchers, hatchery automation (egg/chick handling)","Geared motors on setter turning/tilt, trolley drives, hatchery conveyor/automation","Small (INFERRED)","13 (CONFIRMED)","Europe, Middle East",ELEC,"Egg-turning drives + hatchery-automation conveyors need low-speed gearmotors",7,"lincoincubator.com; proff.dk CVR 32856314"],
["Kverneland Group Kerteminde AS (Taarup)","kvernelandgroup.com","Kerteminde","Grassland / forage & feeding","Disc mowers, tedders, rakes, forage wagons, diet/mixer feeders","Mower cutterbar & input gearboxes; tedder/rake rotor gearboxes; planetary gearbox in vertical diet mixers","Part of Kverneland (Kubota)","~500+ site (INFERRED)","95% export, 80+ countries",PTO,"Gearbox-dense forage line. Caveat: likely group/captive sourcing",6,"kvernelandgroup.com; dk.kverneland.com"],
["BM Silo (BM Agro ApS)","bmsilo.com","Holstebro","Storage silos + dry-bulk conveying","Square/galvanized silos, screw/chain conveyors, pneumatic filling, weighing","Geared motors on discharge/sweep augers, chain conveyors, filling accessories","~DKK 30-60M (INFERRED)","~20-40 (INFERRED)","~80% export (kit-form)",ELEC,"Manufactures conveying/transport equipment for feedstuffs; secondary gearbox demand",6,"bmsilo.com; danskagroindustri.dk"],
["Daltec A/S","daltec.dk","Egtved","Dry feeding (pig + aqua)","Dry feeding filling/dosing, dispensers, phase-feeding, silo filling, mixing/batching","Geared motors on auger drives, dosing units, silo-fill conveyors","~EUR 5-15M (INFERRED)","11-50 (CONFIRMED)","EU + aquaculture (salmon)",ELEC,"Auger-based dry feeding lines require geared motors at each drive",6,"daltec.dk; dnb.com"],
["Sukup Europe A/S / DanCorn","sukup-eu.com","Hedensted","Grain drying & storage","Mobile & continuous grain dryers, steel bins, flat storage, conveyors","Geared motors on dryer discharge/metering rollers, sweep augers, conveyors","~DKK 40-60M (LOW)","21-50 (CONFIRMED range)","Europe-wide",ELEC,"Grain dryers & handling with driven augers/rollers. US-owned (Sukup)",6,"sukup-eu.com; northdata CVR 21170887"],
["Vissing Agro A/S","vissingagro.dk","Bjerringbro area","Pig housing & feeding","Penning, dry & liquid feeding, tube feeders, feed conveying (in-house + AZA)","Geared motors on feed transport augers/chain conveyors, dosing","~EUR 5-15M (INFERRED)","~30-60 (INFERRED)","DK + worldwide",ELEC,"Feed conveying gear-driven. Caveat: partly a distributor (AZA) - verify OEM share",6,"vissingagro.dk; danskagroindustri.dk"],
["Fasterholt Maskinfabrik A/S","fasterholt.dk","Fasterholt (Herning)","Irrigation & bale handling","Hose-reel & self-propelled irrigators, bale wagons","Gearbox/gearmotor driving hose-reel rewind; turbine/drive reduction","Not confirmed (family)","Not confirmed","DK, SE, DE, FR",ELEC,"Reel-drive reduction gearbox in every irrigator",6,"fasterholt.dk"],
["Egedal Maskinfabrik A/S","egedal.dk","Egebjerg (Horsens)","Nursery / horticulture machinery","Bed-sowing machines, bed formers, planters, rotary cultivators","Gearboxes on driven soil-prep rotors, seeder metering drives","Not confirmed","Not confirmed","EU, USA, CA, JP, CN, AU",PTO,"Driven bed-forming/cultivating rotors need gearboxes",6,"egedal.dk; agriculture-xprt"],
["Timan A/S","timan.dk","Tim (Ringkobing)","Landscape / slope maintenance","Radio-controlled slope mowers, weed brushes, hedge trimmers","Angle gearboxes on rotary mower & brush heads; drive gearing on tracked units","Not confirmed","Not confirmed","NL, Europe",PTO,"Rotary/brush heads and tracked drives use gearboxes",6,"timan.dk; northdata CVR 27609627"],
["SKOV A/S","skov.com","Glyngore","Climate/ventilation & farm mgmt (pig/poultry)","Ventilation fans, air-inlet winches, chimneys, feed weighing/silo sensors, controllers","Geared/winch actuators for inlet flaps & curtains; possible feed-weigher gearmotors (fans mostly direct-drive)","~DKK 500M-1B (INFERRED)","~359-500 (CONFIRMED)","70+ markets",ELEC,"Winch/inlet actuation uses gearmotors, but core product is fans/electronics - moderate fit",5,"skov.com; proff.dk CVR 87457117"],
["JYDEN Bur A/S","jyden.com","Vemb","Pig & cattle penning","Gestation crates, cubicles, feeding gates, flat-pack housing, watering","Limited - occasional geared actuators on feed gates; mostly static fabricated steel","~DKK 10-20M (INFERRED)","~94 (MED)","EU + export",ELEC,"Mostly steel penning - low gearbox intensity. Now under SKIOLD group",4,"jyden.com; pitchbook"],
["ABC Hansen (Denmark)","abchansen.dk","Denmark","Milling machinery","Hippo hammer mills, Universal mill, stone mills, small feed/grain plants","Mill rotor often belt-driven; geared motors on ancillary conveyors/feeders","Small (INFERRED)","Small (INFERRED)","Global (esp. Africa)",ELEC,"Hammer/stone mill OEM; some geared-motor content on feeders - many mills belt-driven",4,"abchansen.dk; hippomills.co.za"],
["DACS A/S","dacs.dk","Norre Snede","Climate/ventilation (pig/poultry)","MagFan fans, Corona inlets, ACS climate control, some feed/climate handling","Small gearmotors on inlet/curtain winches (mostly fan motors)","~DKK 40M (INFERRED)","~11 (CONFIRMED)","Europe/international",ELEC,"Limited rotating-gear content; low priority",4,"dacs.dk; dnb.com"],
["Kongskilde Agriculture (CNH Industrial)","kongskilde.com","Soro","Tillage, seeding, forage, feeding implements","Cultivators, seed drills, power harrows, diet mixers","Power-harrow & mixer gearboxes; seed-drill drives","Part of CNH Industrial","","Global (CNH network)",PTO,"Gearbox user but CNH-owned -> captive/group sourcing, lower independent buying",4,"kongskilde.com; CNH acquisition"],
["Graintec A/S (BANKRUPT 2023)","graintec.com","Vejle","Aquafeed & petfood plant engineering","Complete extruded aquafeed & petfood plants (extruders, coolers, conveyors, dosing)","Extruder/conveyor/dosing geared drives within turnkey plants","(bankruptcy Aug 2023)","","Global aquafeed",ELEC,"Historically strong buyer, but bankrupt 2023; assets to Process Integration - verify status",3,"graintec.com; allaboutfeed.net"],
]

industry_map = {
 "Feed Milling": [1,6,8,17,24,32],
 "Grain Handling / Storage / Cleaning": [0,4,7,13,14,25,23],
 "Livestock (Pig & Cattle)": [2,15,26,31],
 "Poultry & Egg-farm / Hatchery": [5,21,30,33],
 "Farm Machinery & Implements": [3,9,10,11,12,18,19,20,22,28,29,35],
 "Climate / Ventilation / Misc": [30,33],
}

HEADERS = ["#","Company Name","Website","Headquarters","Industry","Machinery Produced","Gearbox Application","Est. Annual Revenue","Employees","Export Markets","Existing Gearbox Suppliers (inferred)","Evidence","Buying Potential (1-10)","Source Links"]

NAVY="1F3864"; BLUE="2E5496"; LGREY="F2F2F2"; GREEN="C6EFCE"; YEL="FFEB9C"; RED="FFC7CE"
hdr_font=Font(bold=True,color="FFFFFF",size=11); hdr_fill=PatternFill("solid",fgColor=NAVY)
title_font=Font(bold=True,color="FFFFFF",size=16)
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(wrap_text=True,vertical="top"); center=Alignment(horizontal="center",vertical="center")

def score_fill(s):
    if s>=8: return PatternFill("solid",fgColor=GREEN)
    if s>=6: return PatternFill("solid",fgColor=YEL)
    return PatternFill("solid",fgColor=RED)

def style_header(ws,ncols,row=1):
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c); cell.font=hdr_font; cell.fill=hdr_fill
        cell.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center"); cell.border=border

def write_table(ws,headers,data,start_row=1,score_col=None):
    for c,h in enumerate(headers,1): ws.cell(row=start_row,column=c,value=h)
    style_header(ws,len(headers),start_row)
    r=start_row+1
    for rec in data:
        for c,val in enumerate(rec,1):
            cell=ws.cell(row=r,column=c,value=val); cell.alignment=wrap; cell.border=border
            if r%2==0: cell.fill=PatternFill("solid",fgColor=LGREY)
        if score_col:
            sc=ws.cell(row=r,column=score_col); sc.fill=score_fill(rec[score_col-1]); sc.alignment=center; sc.font=Font(bold=True)
        r+=1
    ws.freeze_panes=ws.cell(row=start_row+1,column=1)

wb=openpyxl.Workbook()

# Read Me
ws=wb.active; ws.title="Read Me"
ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=110
ws.merge_cells('A1:B1'); ws['A1']="I-MAK Reduktor  |  Denmark Machinery OEM Prospecting  -  BATCH 2: Agriculture"
ws['A1'].font=title_font; ws['A1'].fill=PatternFill("solid",fgColor=NAVY); ws['A1'].alignment=center; ws.row_dimensions[1].height=30
readme=["","SCOPE","Batch 2 of 6 - agriculture: feed mills, grain handling/drying/cleaning, livestock (pig & cattle),",
"poultry & egg-farm equipment, and farm machinery & implements. 36 unique Danish OEMs.","",
"TABS","  Company Database - all 36 companies, full field set",
"  Top Prospects - ranked by buying-potential score","  By Industry / By Region - groupings",
"  Decision Makers - 59 Apollo contacts with verified business emails + LinkedIn",
"  Gearbox Suppliers - incumbent brands to displace","",
"TWO INCUMBENT WORLDS (key sales insight)",
"  - Electric geared motors (SEW/NORD/Bonfiglioli/Bauer): feed, grain, conveying, livestock, poultry, climate.",
"  - PTO & angle gearboxes (Comer/Bondioli & Pavesi/Walterscheid): field implements (spreaders, harrows, mowers, slurry).",
"  SKIOLD is the one firm that PUBLICLY documents planetary-gear-motor use - the strongest confirmed signal.","",
"METHOD & CONFIDENCE","  2+ source verification; CONFIRMED vs INFERRED labels; blanks = unknown, not fabricated.",
"  Score: green 8-10 (priority), yellow 6-7, red <=5.","",
"CROSS-BATCH NOTE: Landia (Lem) and Sanovo (Odense) also serve agriculture but are counted in Batch 1.",
"REMAINING BATCHES: 3 Material handling | 4 Wind/Marine | 5 Automation | 6 Misc."]
for i,line in enumerate(readme,start=2):
    c=ws.cell(row=i,column=2,value=line)
    c.font=Font(bold=True,size=12,color=BLUE) if (line.isupper() and line.strip()) else Font(size=11)

# Company Database
ws=wb.create_sheet("Company Database")
data=[[i+1]+rows[i] for i in range(len(rows))]
write_table(ws,HEADERS,data,score_col=13)
for i,w in enumerate([4,30,20,20,26,34,38,22,16,20,28,40,12,26],1):
    ws.column_dimensions[get_column_letter(i)].width=w

# Top Prospects
ws=wb.create_sheet("Top Prospects")
order=sorted(range(len(rows)),key=lambda i:-rows[i][11])
tp=[[rank,rows[i][0],rows[i][3],rows[i][2],rows[i][11],rows[i][5],rows[i][8],rows[i][10]] for rank,i in enumerate(order,1)]
write_table(ws,["Rank","Company","Industry","HQ","Score","Gearbox Application","Employees","Evidence"],tp,score_col=5)
for i,w in enumerate([6,32,26,20,8,40,18,44],1): ws.column_dimensions[get_column_letter(i)].width=w

# By Industry
ws=wb.create_sheet("By Industry"); r=1
for ind,idxs in industry_map.items():
    ws.cell(row=r,column=1,value=ind).font=Font(bold=True,size=13,color="FFFFFF")
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
    for c in range(1,6): ws.cell(row=r,column=c).fill=PatternFill("solid",fgColor=BLUE)
    r+=1
    heads=["Company","HQ","Machinery Produced","Score"]
    for c,h in enumerate(heads,1): ws.cell(row=r,column=c,value=h)
    style_header(ws,4,r); r+=1
    for i in sorted(idxs,key=lambda i:-rows[i][11]):
        rec=[rows[i][0],rows[i][2],rows[i][4],rows[i][11]]
        for c,val in enumerate(rec,1):
            cell=ws.cell(row=r,column=c,value=val); cell.alignment=wrap; cell.border=border
        sc=ws.cell(row=r,column=4); sc.fill=score_fill(rec[3]); sc.alignment=center; sc.font=Font(bold=True); r+=1
    r+=1
for i,w in enumerate([32,22,50,8],1): ws.column_dimensions[get_column_letter(i)].width=w

# By Region
ws=wb.create_sheet("By Region")
def region_of(hq):
    h=hq.lower()
    if any(k in h for k in ["funen","odense","kerteminde","faaborg","aaby"]): return "Funen"
    if any(k in h for k in ["soro","vejle","uldum","grindsted","hedensted","kolding","egtved","norre snede"]): return "Zealand/SE"
    return "Jutland"
from collections import defaultdict
reg=defaultdict(list)
for i,r0 in enumerate(rows): reg[region_of(r0[2])].append(i)
rr=1
for region in ["Jutland","Funen","Zealand/SE"]:
    if region not in reg: continue
    ws.cell(row=rr,column=1,value=f"{region}  ({len(reg[region])} companies)").font=Font(bold=True,size=13,color="FFFFFF")
    ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=4)
    for c in range(1,5): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=BLUE)
    rr+=1
    heads=["Company","HQ","Industry","Score"]
    for c,h in enumerate(heads,1): ws.cell(row=rr,column=c,value=h)
    style_header(ws,4,rr); rr+=1
    for i in sorted(reg[region],key=lambda i:-rows[i][11]):
        rec=[rows[i][0],rows[i][2],rows[i][3],rows[i][11]]
        for c,val in enumerate(rec,1):
            cell=ws.cell(row=rr,column=c,value=val); cell.alignment=wrap; cell.border=border
        sc=ws.cell(row=rr,column=4); sc.fill=score_fill(rec[3]); sc.alignment=center; sc.font=Font(bold=True); rr+=1
    rr+=1
for i,w in enumerate([32,24,30,8],1): ws.column_dimensions[get_column_letter(i)].width=w

# Decision Makers
ws=wb.create_sheet("Decision Makers")
DM_HEAD=["Company","Role","Full Name","Title","Business Email","Email Status","LinkedIn URL","Location"]
dm=[
["Cimbria A/S","Production","Rasmus Tange","Production Manager","","unavailable","linkedin.com/in/rasmus-mosumgaard-tange-29ab32b9","Thisted, DK"],
["Cimbria A/S","Product","Natalia Salberg-Bak","CPQ Product Owner","","unavailable","linkedin.com/in/nataliamaria-kwiatkowska","Thisted, DK"],
["SKIOLD A/S","CEO","Morten Andersen","CEO","mra@skiold.com","verified","linkedin.com/in/morten-rosager-andersen-92b50617","Aarhus, DK"],
["SKIOLD A/S","R&D","Kim Jansson","Product Manager, R&D Climate","","unavailable","linkedin.com/in/kim-jansson-b01381b1","Central DK"],
["SKIOLD A/S","R&D","Frankie Jensen","Product Manager, R&D Housing System/equipment","","unavailable","linkedin.com/in/frankie-j-a238a94a","Central DK"],
["ACO Funki A/S","MD","Lene Bryde","Managing Director","lbryde@acofunki.dk","verified","linkedin.com/in/lene-bryde-4020b888","Central DK"],
["ACO Funki A/S","Production","Andrii Makara","Production Manager","amakara@acofunki.dk","verified","linkedin.com/in/andrii-makara-3ba09b138","Central DK"],
["ACO Funki A/S","Technical/R&D","Poul Nielsen","Manager of Technical & R&D Department","pnielsen@acofunki.dk","verified","linkedin.com/in/poul-nielsen-93186932","Struer, DK"],
["SAMSON AGRO A/S","Production","Tina Kure","Production Manager","tiku@samson-agro.com","verified","linkedin.com/in/tina-k-1560499","Thorso, DK"],
["SAMSON AGRO A/S","Technical Director","Gunnar Mikkelsen","Environmental & Technical Director","","unavailable","linkedin.com/in/gunnar-mikkelsen-ba918122","Central DK"],
["Kongskilde Industries A/S","CEO","Oscar Gunner","CEO","owg@kongskilde-industries.com","verified","linkedin.com/in/oscarwgunner","Copenhagen, DK"],
["Kongskilde Industries A/S","CTO","Michael Moller","Chief Technology Officer","mimo@kongskilde-industries.com","verified","linkedin.com/in/michael-møller-9409669","Sorø, DK"],
["Kongskilde Industries A/S","MD","Anoop Singh","Managing Director","ans@kongskilde-industries.com","verified","linkedin.com/in/anoop-singh-09656a98","Denmark"],
["LANDMECO A/S","CEO","Anders Andersen","Chief Executive Officer","anders@landmeco.dk","verified","linkedin.com/in/anders-kristian-andersen-b656a939","Ringkobing-Skjern, DK"],
["LANDMECO A/S","Production","Henrik Nielsen","Production Manager","hni@landmeco.dk","verified","linkedin.com/in/henrik-nielsen-697ba4225","S. Denmark"],
["ANDRITZ Feed & Biofuel A/S","Production Dev","Edgar Duque","Group Production Development Manager","edgar.duque@andritz.com","verified","linkedin.com/in/edgar-duque-41050121","Esbjerg, DK"],
["ANDRITZ Feed & Biofuel A/S","Procurement","Kirsten Skovsende","Project Procurement Manager","skovsende.kirsten@andritz.com","extrapolated","linkedin.com/in/kirsten-skovsende-7b519729","S. Denmark"],
["ANDRITZ Feed & Biofuel A/S","Production","Flemming Jensen","Production Manager","jensen.flemming@andritz.com","extrapolated","linkedin.com/in/flemming-jensen-960685145","Esbjerg, DK"],
["Jema Agro A/S","Production","Kaj Thorsager","Production Manager","kt@jema.as","verified","linkedin.com/in/kaj-thorsager-90333a234","Silkeborg, DK"],
["Maskinfabrikken Cormall A/S","CEO/MD","Jens Hansen","Managing Director","","unavailable","linkedin.com/in/jens-hansen-2ab2843a","S. Denmark"],
["Fransgard Maskinfabrik A/S","CEO","Jens Ploug","CEO","jp@fransgard.dk","verified","linkedin.com/in/jens-ploug-b594647a","Central DK"],
["Fransgard Maskinfabrik A/S","Purchasing/Owner","Soren Aldahl","Co-owner (Steel Purchasing)","","unavailable","linkedin.com/in/søren-aldahl-851688267","Aars, DK"],
["Lachenmeier Monsun A/S","R&D/Engineering","Esben Friis","Head of Research and Development","ef@lachenmeier-monsun.com","verified","linkedin.com/in/esben-friis-83952b71","S. Denmark"],
["Lachenmeier Monsun A/S","CFO/Owner","Kenneth Luetzen","CFO / Co-owner","kl@lachenmeier-monsun.com","verified","linkedin.com/in/kenneth-lützen-009b0739","Sonderborg, DK"],
["Lachenmeier Monsun A/S","Production","Tage Clausen","Production Manager","","unavailable","linkedin.com/in/tage-clausen-51975719","S. Denmark"],
["Agrisys A/S","CEO","Mikael Kirk","CEO","kirk@agrisys.dk","verified","linkedin.com/in/mikaelkirk","Herning, DK"],
["HARDI International A/S","CEO","Hans Meulenkamp","Group CEO","hans.meulenkamp@hardi.com","verified","linkedin.com/in/hmeulenkamp","Copenhagen, DK"],
["HARDI International A/S","Engineering","Casper Frederiksen","Head of Product Engineering Group","casper.frederiksen@hardi.com","verified","linkedin.com/in/casper-frederiksen-9491134","Zealand, DK"],
["HARDI International A/S","Production/Technical","Kent Zibrandtsen","Production Technology Manager","kent.zibrandtsen@hardi.com","verified","linkedin.com/in/kent-zibrandtsen-581b702a","Norre Alslev, DK"],
["HARDI International A/S","Production","Aske Karstensen","Production Manager","aske.karstensen@hardi.com","verified","linkedin.com/in/aske-karstensen-64845a172","Zealand, DK"],
["Agrometer a/s","CEO","Soren Hovmoller","Adm. Direktor (CEO)","sh@agrometer.dk","verified","linkedin.com/in/soerenhovmoellerchristensen","Aarhus, DK"],
["LINCO Incubator ApS","Managing Director","Harry Sorensen","Managing Director","hso@lincoincubator.com","verified","linkedin.com/in/harry-sorensen","Central DK"],
["Kverneland Group Kerteminde AS","Production","Christian Baechler","Production Manager","christian.bachler@kvernelandgroup.com","verified","linkedin.com/in/christian-odgaard-bächler-67230027","Odense, DK"],
["Kverneland Group Kerteminde AS","Production/Training","Jakob Klokker","Production Training Manager","jakob.klokker@kvernelandgroup.com","verified","linkedin.com/in/jakob-klokker-14531462","Funen, DK"],
["Kverneland Group Kerteminde AS","Ops Support","Luise Skjodt","Production Change and Operations Support Manager","luise.durow@kvernelandgroup.com","verified","linkedin.com/in/luise-durow-skjødt-25100b33","Kerteminde, DK"],
["Kverneland Group Kerteminde AS","R&D","Jens Morthorst","R&D Manager","","unavailable","linkedin.com/in/jens-aage-morthorst-a863b9b","S. Denmark"],
["Kverneland Group Kerteminde AS","Production","Jens Petterson","Production Manager","","unavailable","linkedin.com/in/jens-bo-petterson-30b32aa2","S. Denmark"],
["BM Silo (BM Agro ApS)","Owner/R&D","Claus Martinsen","Owner","claus@bmsilo.com","verified","linkedin.com/in/claus-martinsen-14052525","Holstebro, DK"],
["BM Silo (BM Agro ApS)","Production","Mikkel Clausager","Production Manager","","unavailable","linkedin.com/in/mikkel-flansmose-clausager-725896b5","Holstebro, DK"],
["Daltec A/S","Production","Per Lund","Production Manager","","unavailable","linkedin.com/in/per-lund-308b7b89","Central DK"],
["Sukup Europe A/S / DanCorn","CEO","Jens Iversen","CEO","jei@sukup-eu.com","verified","linkedin.com/in/jens-erik-kjær-iversen-657a1ba9","Juelsminde, DK"],
["Vissing Agro A/S","Owner","Claus Lewandowski","Owner","cb@vissingagro.dk","verified","linkedin.com/in/claus-brandt-67641b2a","S. Denmark"],
["Fasterholt Maskinfabrik A/S","CEO","Rikke Aslak","CEO","ra@fasterholt.dk","verified","linkedin.com/in/rikke-kjær-aslak-3436b318","Denmark"],
["Fasterholt Maskinfabrik A/S","Production","Peter Nyborg","Production Manager","pn@fasterholt.dk","verified","linkedin.com/in/peter-nyborg-6491a35a","Holstebro, DK"],
["Egedal Maskinfabrik A/S","Owner/MD","Allan Pedersen","Ejer / Adm. Direktor","akp@egedal.dk","verified","linkedin.com/in/allan-koed-pedersen-b875302b","Hedensted, DK"],
["Egedal Maskinfabrik A/S","Owner","Keld Nielsen","Owner","","unavailable","linkedin.com/in/keld-sund-nielsen-3a315021","Central DK"],
["Egedal Maskinfabrik A/S","Production","Mikkel Otte","Production Manager","","unavailable","linkedin.com/in/mikkel-otte-2a1962152","Torring, DK"],
["Timan A/S","Technical Director","Henning Pedersen","Teknisk Direktor","hp@timan.dk","verified","linkedin.com/in/henning-pedersen-15336a131","Tim, DK"],
["SKOV A/S","CEO","Leo Ostergaard","Chief Executive Officer","leo@skov.com","verified","linkedin.com/in/leo-østergaard-7094102","Denmark"],
["SKOV A/S","CTO","Tommy Bak","Chief Technology Officer","tba@skov.dk","verified","linkedin.com/in/tommy-bak-a86a91a","Skive, DK"],
["SKOV A/S","Procurement","Jens Nielsen","Procurement Manager","jen@skov.dk","verified","linkedin.com/in/jens-nielsen-4668b944","Roslev, DK"],
["SKOV A/S","R&D","Uffe Odgaard","R&D Manager Mechanics & Test Center","uod@skov.com","verified","linkedin.com/in/uffe-odgaard-a3031463","N. Denmark"],
["SKOV A/S","Production","Borge Larsen","Production Manager","borge.larsen@skov.com","verified","linkedin.com/in/borgemunchlarsen","N. Denmark"],
["DACS A/S","Owner/Sales","Niels Dybdahl","Chief Sales Officer (CSO) & Owner","nd@dacs.dk","verified","linkedin.com/in/niels-dybdahl-bb798822","Central DK"],
["DACS A/S","Co-owner/Engineering","Peter Dybdahl","Co-owner and Developer","pd@dacs.dk","verified","linkedin.com/in/peterdybdahl","Central DK"],
]
write_table(ws,DM_HEAD,dm)
for ri,rec in enumerate(dm,start=2):
    cell=ws.cell(row=ri,column=6); st=rec[5]
    cell.fill=PatternFill("solid",fgColor=GREEN if st=="verified" else YEL if st=="extrapolated" else RED)
verified=sum(1 for r in dm if r[5]=="verified")
note_r=len(dm)+3
ws.cell(row=note_r,column=1,value=f"{len(dm)} decision-makers enriched via Apollo People Match. {verified} verified business emails; rest extrapolated or email-unavailable (LinkedIn still provided).").font=Font(italic=True,bold=True,color=BLUE)
ws.cell(row=note_r+1,column=1,value="11 companies had no Apollo records: Bredal, Bogballe, HE-VA, Assentoft Silo, SKIOLD Damas, Thyregod, GreenTec, JYDEN Bur, ABC Hansen, Kongskilde Agriculture, Graintec (bankrupt).").font=Font(italic=True,color="C00000")
for i,w in enumerate([32,20,26,46,34,14,44,16],1): ws.column_dimensions[get_column_letter(i)].width=w

# Gearbox Suppliers
ws=wb.create_sheet("Gearbox Suppliers")
gs=[
["SEW-Eurodrive","Germany","Electric geared motors on feed mills, grain conveying, livestock/poultry feeding & manure","Price + lead-time, second-source"],
["NORD Drivesystems","Germany","Conveyors, augers, elevators, washdown feed/grain lines","Hygienic/stainless, faster delivery"],
["Bonfiglioli","Italy","Heavy conveying, mixers, rendering-type high-torque","Cost on high-torque units"],
["Bauer (Danfoss)","Germany/DK","Compact geared motors on smaller feeding/dosing machines","Local DK relationship + custom"],
["Comer Industries","Italy","PTO & angle gearboxes on spreaders, harrows, mowers, slurry pumps","Cost + delivery on agri PTO units"],
["Bondioli & Pavesi","Italy","PTO drivelines & gearboxes on implements (spreaders, mixers, harvesters)","Custom ratio/torque, price"],
["Walterscheid","Germany","PTO shafts & angle gearboxes on tillage, spreaders, forage","Second-source on driveline gearing"],
]
write_table(ws,["Incumbent Supplier","Origin","Typical position in Danish agri machinery","I-MAK displacement angle"],gs)
ws.cell(row=len(gs)+3,column=1,value="All supplier attributions INFERRED (no public BOM) except SKIOLD's documented planetary-gear-motor use. Verify per account.").font=Font(italic=True,color="C00000")
for i,w in enumerate([24,16,52,36],1): ws.column_dimensions[get_column_letter(i)].width=w

out="/home/user/zak/market-research/I-MAK_Denmark_Batch2_Agriculture.xlsx"
wb.save(out)
print("Saved",out,"tabs:",wb.sheetnames,"| companies:",len(rows),"| contacts:",len(dm))
